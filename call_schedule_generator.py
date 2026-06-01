from google.colab import drive
from google.colab import auth
import google.auth
from google.auth.transport.requests import Request

# This connects the notebook to your Drive files
drive.mount('/content/drive')

# This authenticates your account to read Google Sheets specifically
auth.authenticate_user()

!pip install ortools gspread

# @title Input Settings
folder_id = "INSERT_GOOGLE_DRIVE_FOLDER_ID" # @param {type:"string"}
output_file_name = "Compiled_Resident_Requests.csv" # @param {type:"string"}

print(f"Targeting folder: {folder_id}")

import pandas as pd
import gspread
from google.auth import default
from google.colab import auth
from googleapiclient.discovery import build
import os
from datetime import timedelta

# 1. Auth and Setup
auth.authenticate_user()
creds, _ = default()
gc = gspread.authorize(creds)
drive_service = build('drive', 'v3')

def get_resident_data(target_folder):
    query = f"'{target_folder}' in parents and mimeType = 'application/vnd.google-apps.spreadsheet'"
    results = drive_service.files().list(q=query, fields="files(id, name)").execute()
    items = results.get('files', [])

    if not items:
        print("No Google Sheets found. Check your Folder ID.")
        return pd.DataFrame()

    all_rows = []

    for item in items:
        resident_name = item['name'].replace('.gsheet', '')
        print(f"Reading: {resident_name}...")

        try:
            sheet = gc.open_by_key(item['id']).get_worksheet(0)
            # Fetch all values as a list of lists to avoid header duplicate errors
            raw_data = sheet.get_all_values()

            if len(raw_data) < 2:
                print(f"  ! {resident_name} is empty.")
                continue

            # Manually find column indices by searching the first row
            header_row = [str(h).lower().strip() for h in raw_data[0]]

            try:
                idx_type = header_row.index('type of request')
                idx_date = header_row.index('date')
                # Priority is optional, default to idx -1 if not found
                idx_prio = header_row.index('priority') if 'priority' in header_row else -1
            except ValueError as e:
                print(f"  ! Missing required column in {resident_name}: {e}")
                continue

            resident_list = []
            # Start from index 1 to skip the header row
            for row in raw_data[1:]:
                date_val = row[idx_date].strip()
                if not date_val: continue

                # Basic cleaning of date and type
                clean_date = pd.to_datetime(date_val, errors='coerce')
                if pd.isna(clean_date): continue

                resident_list.append({
                    'resident_name': resident_name,
                    'type of request': row[idx_type],
                    'date': clean_date,
                    'priority': row[idx_prio] if idx_prio != -1 else 'Low'
                })

            if not resident_list:
                continue

            resident_df = pd.DataFrame(resident_list)

            # Apply the logic rules (Sandwich & Whole Week)
            expanded_df = apply_blackout_rules(resident_df, resident_name)
            all_rows.append(expanded_df)

        except Exception as e:
            print(f"  ! Error processing {resident_name}: {e}")

    if all_rows:
        return pd.concat(all_rows, ignore_index=True)
    return pd.DataFrame()

def apply_blackout_rules(df, name):
    vacation_mask = df['type of request'].str.contains('vacation', case=False, na=False)
    vacation_dates = set(df[vacation_mask]['date'])
    new_blackouts = []
    sorted_dates = sorted(list(vacation_dates))

    for d in sorted_dates:
        # RULE: WHOLE WEEK OFF (Mon-Fri)
        if d.weekday() == 0:
            week_days = [d + timedelta(days=i) for i in range(1, 5)]
            if all(wd in vacation_dates for wd in week_days):
                for offset in [-2, -1, 5, 6]:
                    new_blackouts.append({'date': d + timedelta(days=offset), 'type of request': 'Rule-Based Blackout (Whole Week)'})

        # RULE: SANDWICH RULE (Thu/Fri + Mon/Tue)
        if d.weekday() == 3:
            sandwich_days = [d + timedelta(days=1), d + timedelta(days=4), d + timedelta(days=5)]
            if all(sd in vacation_dates for sd in sandwich_days):
                new_blackouts.append({'date': d + timedelta(days=2), 'type of request': 'Rule-Based Blackout (Sandwich)'})
                new_blackouts.append({'date': d + timedelta(days=3), 'type of request': 'Rule-Based Blackout (Sandwich)'})

    if new_blackouts:
        blackout_df = pd.DataFrame(new_blackouts)
        blackout_df['resident_name'] = name
        blackout_df['priority'] = 'High'
        df = pd.concat([df, blackout_df], ignore_index=True)

    return df.drop_duplicates(subset=['resident_name', 'date'])

# Execute
if 'folder_id' in locals():
    master_df = get_resident_data(folder_id)
    if not master_df.empty:
        master_df = master_df.sort_values(['resident_name', 'date'])
        master_df.to_csv(output_file_name, index=False)
        print(f"\nSUCCESS! {output_file_name} created with {len(master_df)} rows.")
    else:
        print("\nNo data found in sheets.")
else:
    print("Error: Run the 'Input Settings' cell first.")

###-------------------------------------Schedule Validation -----------------------
import pandas as pd
from datetime import timedelta

# Assuming master_df is already in memory from the previous step
# If not, you can load it: master_df = pd.read_csv('Compiled_Resident_Requests.csv', parse_dates=['date'])

# Parameters
MAX_OFF_CUTOFF = 10  # Flag any day where 10 or more residents are unavailable
MAX_TOTAL_DAYS_OFF = 25 # Set to your program's limit (e.g., 4 weeks)
MAX_VACATION_DAYS_PER_WEEKDAY = 4 # New parameter for check 4

print("=========================================")
print("      SCHEDULE VALIDATION REPORT         ")
print("=========================================\n")

# Make sure dates are datetime objects
master_df['date'] = pd.to_datetime(master_df['date'])

# Filter for explicit vacation requests for new checks
vacation_requests = master_df[master_df['type of request'].str.contains('vacation', case=False, na=False)].copy()
vacation_requests['weekday'] = vacation_requests['date'].dt.day_name()


# ---------------------------------------------------------
# CHECK 1: "Too Many Residents Off" on a Single Day
# ---------------------------------------------------------
print("--- 1. HIGH DENSITY DAYS (DANGER ZONES) ---")
daily_counts = master_df.groupby('date')['resident_name'].nunique().reset_index(name='num_off')
danger_days = daily_counts[daily_counts['num_off'] >= MAX_OFF_CUTOFF]

if danger_days.empty:
    print("✅ No days exceed the cutoff limit.")
else:
    for _, row in danger_days.iterrows():
        # Tag if it's a weekend for extra context
        day_type = " (WEEKEND)" if row['date'].weekday() >= 5 else ""
        print(f"⚠️ {row['date'].strftime('%Y-%m-%d')} {row['date'].day_name()}{day_type}: {row['num_off']} residents off!")

print("\n")

# ---------------------------------------------------------
# CHECK 2: 3 Consecutive Same-Weekdays Off
# ---------------------------------------------------------
print("--- 2. CONSECUTIVE SAME-WEEKDAY CHECK ---")
# Use vacation_requests for this check as per user's request
vacation_requests['weekday'] = vacation_requests['date'].dt.day_name()
consecutive_flag = False

# Group by resident and weekday (e.g., Resident1's Mondays)
for (resident, weekday), group in vacation_requests.groupby(['resident_name', 'weekday']):
    # Sort dates to check for 7-day gaps
    dates = sorted(group['date'].tolist())

    for i in range(len(dates) - 2):
        # If date 1, date 2, and date 3 are exactly 7 days apart
        if (dates[i+1] - dates[i] == timedelta(days=7)) and (dates[i+2] - dates[i+1] == timedelta(days=7)):
            print(f"🚩 {resident} has 3 consecutive {weekday}s off starting {dates[i].strftime('%Y-%m-%d')}")
            consecutive_flag = True
            break # Flag it once per cluster to avoid spamming the console

if not consecutive_flag:
    print("✅ No residents have 3 consecutive same-weekdays off.")

print("\n")

# ---------------------------------------------------------
# CHECK 3: Total Days Off Exceeding Limits
# ---------------------------------------------------------
print("--- 3. TOTAL BLACKOUT DAYS PER RESIDENT ---")
# Count only actual requests, not the auto-generated rule blackouts (unless you want to count those)
manual_requests = master_df[~master_df['type of request'].str.contains('Rule-Based', na=False)]
total_off = manual_requests.groupby('resident_name').size()

over_limit = total_off[total_off > MAX_TOTAL_DAYS_OFF]

if over_limit.empty:
    print(f"✅ All residents are under the {MAX_TOTAL_DAYS_OFF} day limit.")
else:
    for resident, count in over_limit.items():
        print(f"🛑 {resident} requested {count} days off (Limit is {MAX_TOTAL_DAYS_OFF}).")

print("\n")

# ---------------------------------------------------------
# NEW CHECK 4: More than N vacation days on a particular weekday
# ---------------------------------------------------------
print(f"--- 4. EXCESS VACATION DAYS PER WEEKDAY (>{MAX_VACATION_DAYS_PER_WEEKDAY}) ---")
excess_weekday_vacations = vacation_requests.groupby(['resident_name', 'weekday']).size().reset_index(name='count')
excess_weekday_vacations = excess_weekday_vacations[excess_weekday_vacations['count'] > MAX_VACATION_DAYS_PER_WEEKDAY]

if excess_weekday_vacations.empty:
    print(f"✅ No residents have more than {MAX_VACATION_DAYS_PER_WEEKDAY} vacation days for any single weekday.")
else:
    for _, row in excess_weekday_vacations.iterrows():
        print(f"🛑 {row['resident_name']} has {row['count']} vacation {row['weekday']}s (Limit {MAX_VACATION_DAYS_PER_WEEKDAY}).")

print("\n")

# ---------------------------------------------------------
# NEW CHECK 5: Vacation days during specific restricted date ranges
# ---------------------------------------------------------
print("--- 5. VACATION DAYS IN RESTRICTED PERIODS ---")
restricted_period_flag = False

# First two weeks of July 2026
july_2026_start = pd.to_datetime('2026-07-01')
july_2026_end = pd.to_datetime('2026-07-14')
july_2026_vacations = vacation_requests[
    (vacation_requests['date'] >= july_2026_start) &
    (vacation_requests['date'] <= july_2026_end)
]

if not july_2026_vacations.empty:
    restricted_period_flag = True
    print("🛑 Vacation requests found in the first two weeks of June 2026:")
    for _, row in july_2026_vacations.iterrows():
        print(f"  - {row['resident_name']} on {row['date'].strftime('%Y-%m-%d')}")

# Last two weeks of June 2027
june_2027_start = pd.to_datetime('2027-06-17')
june_2027_end = pd.to_datetime('2027-06-30')
june_2027_vacations = vacation_requests[
    (vacation_requests['date'] >= june_2027_start) &
    (vacation_requests['date'] <= june_2027_end)
]

if not june_2027_vacations.empty:
    restricted_period_flag = True
    print("🛑 Vacation requests found in the last two weeks of July 2027:")
    for _, row in june_2027_vacations.iterrows():
        print(f"  - {row['resident_name']} on {row['date'].strftime('%Y-%m-%d')}")

if not restricted_period_flag:
    print("✅ No vacation days requested during restricted periods.")
print("\n")

# ---------------------------------------------------------
# NEW CHECK 6: Vacation days on PRITE days
# ---------------------------------------------------------
print("--- 6. VACATION DAYS ON PRITE DAYS ---")
prite_days = [pd.to_datetime('2026-09-29'), pd.to_datetime('2026-10-06')]
prite_vacations = vacation_requests[vacation_requests['date'].isin(prite_days)]
prite_flag = False

if not prite_vacations.empty:
    prite_flag = True
    print("🛑 Vacation requests found on PRITE days:")
    for _, row in prite_vacations.iterrows():
        print(f"  - {row['resident_name']} on {row['date'].strftime('%Y-%m-%d')}")

if not prite_flag:
    print("✅ No vacation days requested on PRITE days.")
print("\n")

# ---------------------------------------------------------
# NEW CHECK 7: Vacation days during retreat
# ---------------------------------------------------------
print("--- 7. VACATION DAYS DURING RETREAT ---")
retreat_days = [pd.to_datetime('2026-08-31'), pd.to_datetime('2026-09-01')]
retreat_vacations = vacation_requests[vacation_requests['date'].isin(retreat_days)]
retreat_flag = False

if not retreat_vacations.empty:
    retreat_flag = True
    print("🛑 Vacation requests found during the retreat:")
    for _, row in retreat_vacations.iterrows():
        print(f"  - {row['resident_name']} on {row['date'].strftime('%Y-%m-%d')}")

if not retreat_flag:
    print("✅ No vacation days requested during retreat.")
print("\n")

print("=========================================")
print("             END OF REPORT               ")
print("=========================================")

### Optimization Model Linear Version -------------------------------------
import pandas as pd
import numpy as np
from datetime import date, timedelta
from ortools.sat.python import cp_model
import time # Import time module

# --- USER SETTINGS ---
INPUT_CSV = 'Compiled_Resident_Requests.csv'
OUTPUT_SCHEDULE = 'Final_Optimized_Schedule.csv'
START_DATE = '2026-07-01'
END_DATE = '2027-06-30'
RETREAT_DATE = '2026-08-31' # Added: Specific day with no resident assignment

# Define your major holidays (YYYY-MM-DD)
HOLIDAYS = [
    '2026-07-03', # Independence Day
    '2026-09-07', # Labor Day
    '2026-11-11', # Veterans Day
    '2026-11-26', # Thanksgiving
    '2026-11-27', # Day after Thanksgiving
    '2026-12-24', # Christmas Eve
    '2026-12-25', # Christmas Day
    '2026-12-31', # New Year's Eve
    '2027-01-01', # New Year's Day
    '2027-01-18', # MLK Day
    '2027-02-15', # President's Day
    '2027-03-26', # Farmworker's Day
    '2027-05-31', # Memorial Day
    '2027-06-18'  # Juneteenth Day
]

print("Loading data and setting up the mathematical model...")

# 1. Load Data & Prep Calendar
requests_df = pd.read_csv(INPUT_CSV, parse_dates=['date'])
residents = sorted(requests_df['resident_name'].unique().tolist())
num_residents = len(residents)

all_dates = pd.date_range(start=START_DATE, end=END_DATE)
num_days = len(all_dates)
# Identify retreat index
retreat_idx = all_dates.get_loc(RETREAT_DATE)

holiday_dates = pd.to_datetime(HOLIDAYS)
is_weekend = [1 if d.weekday() >= 5 else 0 for d in all_dates]
is_holiday = [1 if d in holiday_dates else 0 for d in all_dates]

# Convert blackouts into a fast lookup dictionary: blackouts[(resident, date_index)] = True
blackout_dict = {}
for _, row in requests_df.iterrows():
    r_idx = residents.index(row['resident_name'])
    try:
        d_idx = all_dates.get_loc(row['date'])
        blackout_dict[(r_idx, d_idx)] = True
    except KeyError:
        pass # Date is outside the academic year

# 2. Initialize the Model
model = cp_model.CpModel()

# Variables: shifts[(r, d)] is 1 if resident r works on day d, else 0
shifts = {}
for r_idx in range(num_residents):
    for d_idx in range(num_days):
        shifts[(r_idx, d_idx)] = model.NewBoolVar(f'shift_r{r_idx}_d{d_idx}')

# --- HARD CONSTRAINTS ---

# 1. Resident per shift logic
for d_idx in range(num_days):
    if d_idx == retreat_idx:
        # Modified: No one assigned on Retreat Day
        model.Add(sum(shifts[(r_idx, d_idx)] for r_idx in range(num_residents)) == 0)
    else:
        # Exactly one resident per shift
        model.AddExactlyOne([shifts[(r_idx, d_idx)] for r_idx in range(num_residents)])

# 2. No shifts on requested blackout days
for (r_idx, d_idx) in blackout_dict.keys():
    model.Add(shifts[(r_idx, d_idx)] == 0)

# 3. No back-to-back shifts (If working day d, cannot work d+1)
for r_idx in range(num_residents):
    for d_idx in range(num_days - 1):
        model.Add(shifts[(r_idx, d_idx)] + shifts[(r_idx, d_idx + 1)] <= 1)

# --- SOFT CONSTRAINTS (FAIRNESS & DENSITY) ---

# 1. Map days to ISO weeks (Year, Week Number)
# We create a dictionary where keys are (Year, Week) and values are lists of day indices
weeks_map = {}
for d_idx, d_val in enumerate(all_dates):
    yr_wk = d_val.isocalendar()[:2]
    if yr_wk not in weeks_map:
        weeks_map[yr_wk] = []
    weeks_map[yr_wk].append(d_idx)

unique_weeks = sorted(weeks_map.keys())

# Target averages for balancing
target_total = (num_days - 1) // num_residents # Adjusted for retreat day
target_weekend = sum(is_weekend) // num_residents
target_holiday = sum(is_holiday) // num_residents

# Calculate target for weekdays (days that are not weekends or holidays)
target_weekday = (num_days - 1 - sum(is_weekend) - sum(is_holiday)) // num_residents

resident_penalties = []

for r_idx in range(num_residents):
    # --- Part A: Standard Fairness (Totals, Weekends, Holidays) ---
    r_total = sum(shifts[(r_idx, d_idx)] for d_idx in range(num_days))
    r_weekend = sum(shifts[(r_idx, d_idx)] * is_weekend[d_idx] for d_idx in range(num_days))
    r_holiday = sum(shifts[(r_idx, d_idx)] * is_holiday[d_idx] for d_idx in range(num_days))
    r_weekday = r_total - r_weekend - r_holiday # Calculate weekday shifts for resident

    # Calculate absolute differences from average
    diff_weekday = model.NewIntVar(0, num_days, f'diff_weekday_r{r_idx}')
    model.Add(diff_weekday >= r_weekday - target_weekday)
    model.Add(diff_weekday >= target_weekday - r_weekday)

    diff_weekend = model.NewIntVar(0, num_days, f'diff_week_r{r_idx}')
    model.Add(diff_weekend >= r_weekend - target_weekend)
    model.Add(diff_weekend >= target_weekend - r_weekend)

    diff_holiday = model.NewIntVar(0, num_days, f'diff_hol_r{r_idx}')
    model.Add(diff_holiday >= r_holiday - target_holiday)
    model.Add(diff_holiday >= target_holiday - r_holiday)

    # --- Part B: Minimize Shift Density (>1 shift per week) ---
    multi_shift_vars = []
    for yr_wk in unique_weeks:
        day_indices = weeks_map[yr_wk]
        shifts_in_week = [shifts[(r_idx, d_idx)] for d_idx in day_indices]

        # is_multi is 1 if sum of shifts in week > 1, else 0
        is_multi = model.NewBoolVar(f'is_multi_r{r_idx}_wk{yr_wk}')

        # Logic: If sum > 1, then is_multi must be 1
        # Logic: If sum <= 1, then is_multi must be 0
        model.Add(sum(shifts_in_week) > 1).OnlyEnforceIf(is_multi)
        model.Add(sum(shifts_in_week) <= 1).OnlyEnforceIf(is_multi.Not())
        multi_shift_vars.append(is_multi)

    r_multi_count = sum(multi_shift_vars)

    # --- Part C: Linear penalty for total shift volume deviation (integrated below) ---
    # The deviation 'diff_weekday' is used directly in the penalty calculation.

    # --- Part D: Combine into a single resident penalty score ---
    # Weights: Holiday (5x), Weekend (3x), Multi-shift Week (2x), Weekday (linear with SHIFT_VOLUME_WEIGHT)
    res_penalty = model.NewIntVar(0, 1000000, f'res_penalty_{r_idx}') # Max needed is around 4000
    model.Add(res_penalty ==
                             (diff_weekday * 1) +
                             (diff_weekend * 3) +
                             (diff_holiday * 4) +
                             (r_multi_count * 2))

    resident_penalties.append(res_penalty)

# --- THE MINIMAX BALANCER ---
# The maximum possible penalty is roughly (3*365 + 5*365 + 2*52 + 2*365) = ~3754, so 1,000,000 is safe.
max_penalty = model.NewIntVar(0, 1000000, 'max_penalty')
for p in resident_penalties:
    model.Add(max_penalty >= p)

# Minimize the outlier (the person with the worst schedule) and group sum simultaneously
model.Minimize((max_penalty * 100) + sum(resident_penalties))

# 3. Solve the Model
print("Running CP-SAT Optimizer (This may take 10-60 seconds)...")
solver = cp_model.CpSolver()
solver.parameters.max_time_in_seconds = 300.0 # Stop trying after 5 minutes

start_time = time.time() # Start time measurement
status = solver.Solve(model)
end_time = time.time() # End time measurement

if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
    print(f"✅ Solution Found! Status: {solver.StatusName(status)}")
    print(f"Model solved in {end_time - start_time:.2f} seconds") # Print the elapsed time

    # Extract results
    schedule_data = []
    for d_idx in range(num_days):
        assigned_name = "RETREAT" if d_idx == retreat_idx else "NONE"
        assigned_type = "Special" if d_idx == retreat_idx else "Weekday"

        for r_idx in range(num_residents):
            if solver.BooleanValue(shifts[(r_idx, d_idx)]):
                assigned_name = residents[r_idx]
                assigned_type = 'Holiday' if is_holiday[d_idx] else ('Weekend' if is_weekend[d_idx] else 'Weekday')
                break

        schedule_data.append({
            'Date': all_dates[d_idx].strftime('%Y-%m-%d'),
            'Day of Week': all_dates[d_idx].day_name(),
            'Resident Assigned': assigned_name,
            'Shift Type': assigned_type
        })

    pd.DataFrame(schedule_data).to_csv(OUTPUT_SCHEDULE, index=False)

    # Generate Fairness Report
    print("\n========================================================")
    print("                 FAIRNESS AUDIT REPORT                  ")
    print("========================================================")

    fairness_data = []
    for r_idx in range(num_residents):
        total = sum(solver.BooleanValue(shifts[(r_idx, d_idx)]) for d_idx in range(num_days))
        weekend = sum(solver.BooleanValue(shifts[(r_idx, d_idx)]) * is_weekend[d_idx] for d_idx in range(num_days))
        holiday = sum(solver.BooleanValue(shifts[(r_idx, d_idx)]) * is_holiday[d_idx] for d_idx in range(num_days))
        weekday_shifts = total - weekend - holiday

        # --- Shift Density Check ---
        assigned_dates = [all_dates[d_idx] for d_idx in range(num_days) if solver.BooleanValue(shifts[(r_idx, d_idx)])]
        week_counts = {}
        for current_date in assigned_dates:
            yr_wk = current_date.isocalendar()[:2]
            week_counts[yr_wk] = week_counts.get(yr_wk, 0) + 1
        multi_shift_weeks = sum(1 for count in week_counts.values() if count > 1)

        p_score = solver.Value(resident_penalties[r_idx])

        fairness_data.append({
            'Resident': residents[r_idx],
            'Total': total,
            'Weekdays': weekday_shifts,
            'Weekends': weekend,
            'Holidays': holiday,
            '>1 Shift Weeks': multi_shift_weeks,
            'Penalty Score': p_score
        })

    fair_df = pd.DataFrame(fairness_data)
    print(fair_df.to_string(index=False))

    print("\n--- METRICS ---")
    print(f"Target Total Average: ~{target_total} shifts")
    print(f"Target Weekend Average: ~{target_weekend} shifts")
    print(f"Target Holiday Average: ~{target_holiday} shifts")
    print(f"Target Weekday Average: ~{target_weekday} shifts") # Added new metric
    print(f"\nFinal Schedule saved as: {OUTPUT_SCHEDULE}")
    print("========================================================")

elif status == cp_model.INFEASIBLE:
    print("🛑 FAILED: The constraints are Impossible to solve.")
else:
    print("⚠️ FAILED: Solver timed out before finding a solution.")

### Nice Schedule Builder------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
import calendar

def create_calendar_excel(input_csv, output_xlsx, fairness_df):
    df = pd.read_csv(input_csv)
    df['Date'] = pd.to_datetime(df['Date'])

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Styles
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    weekend_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    holiday_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    retreat_fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid") # Added style

    white_font = Font(color="FFFFFF", bold=True, size=12)
    black_bold = Font(bold=True)
    retreat_font = Font(bold=True, color="C00000") # Added font
    date_font = Font(size=10, color="666666")

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)

    months = df['Date'].dt.to_period('M').unique()

    for month_period in sorted(months):
        month_df = df[df['Date'].dt.to_period('M') == month_period].copy()
        sheet_name = month_period.strftime('%B %Y')
        ws = wb.create_sheet(title=sheet_name)

        for col in range(1, 8):
            col_letter = chr(64 + col)
            ws.column_dimensions[col_letter].width = 22

        ws.merge_cells('A1:G1')
        title_cell = ws.cell(row=1, column=1, value=sheet_name.upper())
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = alignment_center

        weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for i, day in enumerate(weekdays):
            cell = ws.cell(row=2, column=i+1, value=day)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = alignment_center

        start_date = month_period.to_timestamp()
        start_col = start_date.weekday()

        for _, row_data in month_df.iterrows():
            d = row_data['Date'].day
            wd = row_data['Date'].weekday()
            col_idx = wd + 1
            week_offset = (d + start_col - 1) // 7
            grid_row = 3 + (week_offset * 2)

            date_cell = ws.cell(row=grid_row, column=col_idx, value=d)
            date_cell.font = date_font
            date_cell.border = thin_border
            date_cell.alignment = Alignment(horizontal='right', vertical='top')

            name_cell = ws.cell(row=grid_row+1, column=col_idx, value=row_data['Resident Assigned'])
            name_cell.border = thin_border
            name_cell.alignment = alignment_center

            # Apply Conditional Shading
            if row_data['Resident Assigned'] == 'RETREAT':
                name_cell.font = retreat_font
                name_cell.fill = retreat_fill
                date_cell.fill = retreat_fill
            else:
                name_cell.font = black_bold
                if row_data['Shift Type'] == 'Weekend':
                    date_cell.fill = weekend_fill
                    name_cell.fill = weekend_fill
                elif row_data['Shift Type'] == 'Holiday':
                    date_cell.fill = holiday_fill
                    name_cell.fill = holiday_fill

        ws.freeze_panes = "A3"

    # NEW FEATURE: Monthly Shift Summary per Resident
    summary_ws = wb.create_sheet(title="Monthly Shifts Summary")
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 15

    residents = df['Resident Assigned'].unique()
    start_row = 1

    for resident in sorted(residents):
        if resident == "RETREAT" or resident == "NONE": # Skip special entries
            continue

        # Filter for the current resident and exclude 'RETREAT' shifts (which are not real shifts)
        resident_shifts_df = df[(df['Resident Assigned'] == resident) & (df['Resident Assigned'] != 'RETREAT')]

        if resident_shifts_df.empty:
            continue

        # Group by month and count shifts
        monthly_shifts = resident_shifts_df.groupby(resident_shifts_df['Date'].dt.to_period('M')).size()

        if monthly_shifts.empty:
            continue

        # Convert to DataFrame for easier handling and add a total row
        monthly_shifts_df = monthly_shifts.reset_index(name='Shifts').rename(columns={'Date': 'Month'})
        monthly_shifts_df['Month'] = monthly_shifts_df['Month'].dt.strftime('%Y-%m')

        # Add a total row
        total_shifts = monthly_shifts_df['Shifts'].sum()
        monthly_shifts_df.loc['Total'] = ['Total', total_shifts]

        # Write resident's name
        name_cell = summary_ws.cell(row=start_row, column=1, value=f"Resident: {resident}")
        name_cell.font = Font(bold=True, size=14)
        start_row += 2 # Leave a gap

        # Write header
        header_cells = ['Month', 'Shifts']
        for col_idx, header in enumerate(header_cells):
            cell = summary_ws.cell(row=start_row, column=col_idx + 1, value=header)
            cell.font = black_bold
            cell.border = thin_border
            cell.alignment = alignment_center
        start_row += 1

        # Write data rows
        for r_idx, row_data in monthly_shifts_df.iterrows():
            for col_idx, value in enumerate(row_data):
                cell = summary_ws.cell(row=start_row, column=col_idx + 1, value=value)
                cell.border = thin_border
                cell.alignment = alignment_center
                if r_idx == 'Total': # Apply bold to total row
                    cell.font = black_bold
            start_row += 1
        start_row += 2 # Gap before next resident

    # Add Fairness Report sheet
    fairness_ws = wb.create_sheet(title="Fairness Report")
    # Write DataFrame headers
    for col_idx, header in enumerate(fairness_df.columns):
        cell = fairness_ws.cell(row=1, column=col_idx + 1, value=header)
        cell.font = black_bold
        cell.border = thin_border
        cell.alignment = alignment_center
    # Write DataFrame data
    for r_idx, row_data in fairness_df.iterrows():
        for col_idx, value in enumerate(row_data):
            cell = fairness_ws.cell(row=r_idx + 2, column=col_idx + 1, value=value)
            cell.border = thin_border
            cell.alignment = alignment_center
    # Adjust column widths
    for col_idx, column in enumerate(fairness_df.columns):
        max_length = 0
        column_values = [column] + fairness_df[column].astype(str).tolist()
        for cell_value in column_values:
            if len(str(cell_value)) > max_length:
                max_length = len(str(cell_value))
        fairness_ws.column_dimensions[chr(65 + col_idx)].width = max_length + 2 # Add a little padding

    wb.save(output_xlsx)
    print(f"Calendar formatted Excel saved as: {output_xlsx}")

create_calendar_excel('Final_Optimized_Schedule.csv', 'Call_Schedule_Visual_Calendar.xlsx', fair_df)